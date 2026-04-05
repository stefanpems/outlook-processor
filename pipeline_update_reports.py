"""
Update XLSX and HTML report files with current session data.
Reads session_state.json and rebuilds both reports.

Usage: python pipeline_update_reports.py
Reads: session_state.json (contains emails list with all status fields)
Writes: the XLSX and HTML files referenced in session_state.json
"""
import json, re, os, sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

sys.stdout = open(sys.stdout.fileno(), mode="w", buffering=1)

BASE = os.path.dirname(os.path.abspath(__file__))
CONFIG = json.load(open(os.path.join(BASE, "config.json"), encoding="utf-8"))
SESSION_FILE = os.path.join(BASE, "session_state.json")

COLOR_PALETTE = CONFIG.get("topic_color_palette", [
    "#F0E6D3", "#D3E8F0", "#D3F0D6", "#F0D3E6", "#E6F0D3", "#D3D8F0",
    "#F0DAD3", "#D3F0EA", "#E8D3F0", "#F0F0D3", "#D3EAF0", "#E6D3F0",
    "#F0D3D3", "#D3F0D3", "#D3D3F0", "#F0ECD3", "#F0D3EC", "#D3F0F0",
])

TOPIC_COLORS = {}


def get_topic_color(topic):
    if topic not in TOPIC_COLORS:
        idx = len(TOPIC_COLORS) % len(COLOR_PALETTE)
        TOPIC_COLORS[topic] = COLOR_PALETTE[idx]
    return TOPIC_COLORS[topic]


def update_xlsx(xlsx_path, emails):
    """Rebuild the XLSX report from current email data."""
    TOPIC_COLORS.clear()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processed Emails"

    HEADERS = [
        "Published", "Title", "Tech", "Link", "Topic", "Summary",
        "Formatted Summary", "Dup Session", "Dup SP",
        "SP Created", "Categorized", "Moved",
    ]
    hfill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sort by topic (asc) then published_date (desc)
    sorted_data = sorted(emails, key=lambda e: (
        e.get("topic", "").lower(),
        "".join(chr(255 - ord(c)) if c.isdigit() else c
                for c in e.get("published_date", "")) if e.get("published_date") else "~",
    ))

    yes_font = Font(bold=True, color="2E7D32")  # Green for Yes
    link_font = Font(color="0066CC", underline="single")

    for row_idx, em in enumerate(sorted_data, 2):
        color_hex = get_topic_color(em.get("topic", "")).lstrip("#")
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        # Published
        pub = em.get("published_date", "").replace("-", ".")
        ws.cell(row=row_idx, column=1, value=pub).fill = fill

        # Title
        ws.cell(row=row_idx, column=2, value=em.get("title", "")).fill = fill

        # Tech
        ws.cell(row=row_idx, column=3, value=em.get("tech", "")).fill = fill

        # Link
        lc = ws.cell(row=row_idx, column=4, value=em.get("blog_link", ""))
        lc.fill = fill
        lc.font = link_font

        # Topic
        ws.cell(row=row_idx, column=5, value=em.get("topic", "")).fill = fill

        # Summary (plain text — strip HTML)
        plain_summary = re.sub(r'<[^>]+>', '', em.get("summary", ""))
        sc = ws.cell(row=row_idx, column=6, value=plain_summary)
        sc.fill = fill
        sc.alignment = Alignment(wrap_text=True, vertical="top")

        # Formatted Summary (with HTML tags)
        fc = ws.cell(row=row_idx, column=7, value=em.get("summary", ""))
        fc.fill = fill
        fc.alignment = Alignment(wrap_text=True, vertical="top")

        # Status columns
        for col_idx, field in [(8, "dup_session"), (9, "dup_sp"),
                               (10, "sp_created"), (11, "categorized"), (12, "moved")]:
            val = em.get(field, "")
            if isinstance(val, bool):
                val = "Yes" if val else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            if val == "Yes":
                cell.font = yes_font

    # Column widths
    widths = {"A": 12, "B": 55, "C": 30, "D": 30, "E": 18, "F": 65,
              "G": 65, "H": 13, "I": 13, "J": 13, "K": 13, "L": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


def update_html(html_path, emails, session_info=None):
    """Rebuild the HTML report with topic sections and article sub-sections."""
    TOPIC_COLORS.clear()

    # Group by topic
    topics = {}
    for em in emails:
        topic = em.get("topic", "Unknown")
        if topic not in topics:
            topics[topic] = []
        topics[topic].append(em)

    # Sort topics alphabetically
    sorted_topics = sorted(topics.keys())

    # Sort articles within each topic by published_date descending
    for topic in sorted_topics:
        topics[topic].sort(key=lambda e: e.get("published_date", ""), reverse=True)

    # Stats
    total = len(emails)
    dup_session = sum(1 for e in emails if e.get("dup_session") == "Yes" or e.get("dup_session") is True)
    dup_sp = sum(1 for e in emails if e.get("dup_sp") == "Yes" or e.get("dup_sp") is True)
    sp_created = sum(1 for e in emails if e.get("sp_created") == "Yes")
    categorized = sum(1 for e in emails if e.get("categorized") == "Yes")
    moved = sum(1 for e in emails if e.get("moved") == "Yes")

    date_range = ""
    if session_info:
        date_range = f"{session_info.get('date_from', '')} to {session_info.get('date_to', '')}"

    # Build HTML
    html = """<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>Blog Notifications""" + (f" — {date_range}" if date_range else "") + """</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
    background: #f0f2f5; color: #1a1a2e;
    line-height: 1.6; padding: 2rem 1rem;
    max-width: 1200px; margin: 0 auto;
  }
  h1 {
    color: #1a1a2e; font-size: 1.8rem;
    border-bottom: 3px solid #4361ee;
    padding-bottom: .5rem; margin-bottom: 1.5rem;
  }
  .stats {
    background: linear-gradient(135deg, #e8eaf6, #f3e5f5);
    padding: 1rem 1.5rem; border-radius: 10px;
    margin-bottom: 2rem; font-size: .95rem;
    box-shadow: 0 2px 6px rgba(0,0,0,.08);
    display: flex; gap: 2rem; flex-wrap: wrap;
  }
  .stats .stat-item { font-weight: 600; }
  .stats .stat-value { color: #4361ee; }
  .toc {
    background: #fff; padding: 1.5rem 2rem; border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,.07); margin-bottom: 2rem;
  }
  .toc h2 { font-size: 1.2rem; margin-bottom: .8rem; color: #4361ee; }
  .toc ul { list-style: none; columns: 2; column-gap: 2rem; }
  .toc li { padding: .35rem 0; }
  .toc a {
    color: #4361ee; text-decoration: none; font-weight: 500;
    transition: color .2s;
  }
  .toc a:hover { color: #3a0ca3; text-decoration: underline; }
  .toc .count { color: #888; font-size: .85rem; margin-left: .3rem; }
  .topic-section {
    background: #fff; border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,.07);
    padding: 1.5rem 2rem; margin-bottom: 1.5rem;
  }
  .topic-header {
    font-size: 1.3rem; color: #1a1a2e;
    padding-left: .5rem; margin-bottom: 1rem;
    display: flex; align-items: center; gap: .5rem;
  }
  .topic-header .badge {
    background: #4361ee; color: #fff; font-size: .75rem;
    padding: 2px 10px; border-radius: 12px;
  }
  .article {
    border-bottom: 1px solid #eee; padding: 1rem 0;
    transition: background .15s;
  }
  .article:last-child { border-bottom: none; }
  .article:hover { background: #fafbff; border-radius: 6px; padding-left: .5rem; }
  .article-title { font-size: 1.05rem; font-weight: 600; }
  .article-title a { color: #4361ee; text-decoration: none; }
  .article-title a:hover { text-decoration: underline; color: #3a0ca3; }
  .article-meta {
    font-size: .85rem; color: #666; margin: .3rem 0 .6rem;
    display: flex; gap: .8rem; flex-wrap: wrap; align-items: center;
  }
  .article-meta .date { font-weight: 500; }
  .tag {
    background: #e8eaf6; padding: 2px 10px; border-radius: 12px;
    font-size: .78rem; color: #4361ee; display: inline-block;
  }
  .article-summary {
    font-size: .93rem; line-height: 1.6; color: #333;
    max-width: 900px;
  }
  .article-summary b { color: #1a1a2e; }
  .article-summary ul { margin: .4rem 0 .4rem 1.2rem; }
  .article-summary li { margin-bottom: .2rem; }
  .back-link {
    display: inline-block; margin-top: 1rem; color: #4361ee;
    font-size: .85rem; text-decoration: none; font-weight: 500;
  }
  .back-link:hover { text-decoration: underline; }
  .dup-badge {
    background: #fff3cd; color: #856404; padding: 2px 8px;
    border-radius: 4px; font-size: .78rem; font-weight: 500;
  }
  footer {
    text-align: center; color: #999; font-size: .8rem;
    margin-top: 3rem; padding-top: 1rem;
    border-top: 1px solid #e0e0e0;
  }
</style></head><body>
"""

    html += f'<h1>Blog Notifications' + (f' &mdash; {date_range}' if date_range else '') + '</h1>\n'

    # Stats bar
    html += '<div class="stats">\n'
    html += f'  <div class="stat-item">Total: <span class="stat-value">{total}</span></div>\n'
    html += f'  <div class="stat-item">Dup Session: <span class="stat-value">{dup_session}</span></div>\n'
    html += f'  <div class="stat-item">Dup SP: <span class="stat-value">{dup_sp}</span></div>\n'
    html += f'  <div class="stat-item">SP Created: <span class="stat-value">{sp_created}</span></div>\n'
    html += f'  <div class="stat-item">Categorized: <span class="stat-value">{categorized}</span></div>\n'
    html += f'  <div class="stat-item">Moved: <span class="stat-value">{moved}</span></div>\n'
    html += '</div>\n'

    # Table of contents
    html += '<nav class="toc" id="toc"><h2>Topics</h2><ul>\n'
    for topic in sorted_topics:
        topic_id = re.sub(r'[^a-zA-Z0-9]', '-', topic).lower()
        count = len(topics[topic])
        html += f'  <li><a href="#{topic_id}">{topic}</a> <span class="count">({count})</span></li>\n'
    html += '</ul></nav>\n'

    # Topic sections
    for topic in sorted_topics:
        topic_id = re.sub(r'[^a-zA-Z0-9]', '-', topic).lower()
        color = get_topic_color(topic)

        html += f'<section class="topic-section" id="{topic_id}" style="border-left-color: {color};">\n'
        html += f'  <h2 class="topic-header">{topic} <span class="badge">{len(topics[topic])}</span></h2>\n'

        for em in topics[topic]:
            link = em.get("blog_link", "")
            title = em.get("title", "")
            pub = em.get("published_date", "").replace("-", ".")
            tech = em.get("tech", "")
            summary = em.get("summary", "")
            is_dup = em.get("dup_session") == "Yes" or em.get("dup_session") is True \
                     or em.get("dup_sp") == "Yes" or em.get("dup_sp") is True

            html += '  <div class="article">\n'
            html += f'    <div class="article-title">'
            if link:
                html += f'<a href="{link}" target="_blank">{title}</a>'
            else:
                html += title
            if is_dup:
                html += ' <span class="dup-badge">Duplicate</span>'
            html += '</div>\n'

            html += '    <div class="article-meta">'
            if pub:
                html += f'<span class="date">{pub}</span>'
            if tech:
                for tag in tech.split(","):
                    tag = tag.strip()
                    if tag:
                        html += f' <span class="tag">{tag}</span>'
            html += '</div>\n'

            if summary:
                html += f'    <div class="article-summary">{summary}</div>\n'

            html += '  </div>\n'

        html += f'  <a href="#toc" class="back-link">&uarr; Back to index</a>\n'
        html += '</section>\n'

    html += '<footer>Generated by Blog Notification Pipeline</footer>\n'
    html += '</body></html>'

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)


def main():
    if not os.path.exists(SESSION_FILE):
        print("Error: session_state.json not found", file=sys.stderr)
        sys.exit(1)

    session = json.load(open(SESSION_FILE, encoding="utf-8"))
    xlsx_path = session.get("xlsx_path", "")
    html_path = session.get("html_path", "")
    emails = session.get("emails", [])

    if not xlsx_path or not html_path:
        print("Error: session_state.json missing xlsx_path or html_path", file=sys.stderr)
        sys.exit(1)

    update_xlsx(xlsx_path, emails)
    update_html(html_path, emails, session)

    print(json.dumps({
        "xlsx_path": xlsx_path,
        "html_path": html_path,
        "total_emails": len(emails),
    }, indent=2))


if __name__ == "__main__":
    main()
