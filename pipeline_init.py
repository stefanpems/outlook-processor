"""
Initialize a blog notification processing session.
Creates XLSX + HTML report template files with progressive naming.

Usage: python pipeline_init.py
Output: JSON to stdout with session info (xlsx_path, html_path, session_file, date).
"""
import json, os, sys, glob
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))
CONFIG = json.load(open(os.path.join(BASE, "config.json"), encoding="utf-8"))
OUTPUT = os.path.join(BASE, CONFIG["output"]["dir"])
os.makedirs(OUTPUT, exist_ok=True)

# ── Determine progressive file name ──────────────────────────────────────
today = datetime.now().strftime("%Y.%m.%d")
nn = 1
while True:
    base_name = f"{today}-{nn:02d}-ProcessedEmails"
    xlsx_path = os.path.join(OUTPUT, f"{base_name}.xlsx")
    html_path = os.path.join(OUTPUT, f"{base_name}.html")
    if not os.path.exists(xlsx_path) and not os.path.exists(html_path):
        break
    nn += 1

session_file = os.path.join(BASE, "session_state.json")

# ── Create XLSX template ─────────────────────────────────────────────────
HEADERS = [
    "Published", "Title", "Tech", "Link", "Topic", "Summary",
    "Formatted Summary", "Dup Session", "Dup SP",
    "SP Created", "Categorized", "Moved",
]
COL_WIDTHS = {
    "A": 12, "B": 55, "C": 30, "D": 30, "E": 18, "F": 65,
    "G": 65, "H": 13, "I": 13, "J": 13, "K": 13, "L": 13,
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Processed Emails"
hfill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
hfont = Font(color="FFFFFF", bold=True, size=10)

for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal="center", vertical="center")

for col_letter, width in COL_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

# Freeze header row
ws.freeze_panes = "A2"
wb.save(xlsx_path)

# ── Create HTML template ─────────────────────────────────────────────────
html = """<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>Blog Notifications — Processing Session {date}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
    background: #f0f2f5; color: #1a1a2e;
    line-height: 1.6; padding: 2rem 1rem;
    max-width: 1200px; margin: 0 auto;
  }}
  h1 {{
    color: #1a1a2e; font-size: 1.8rem;
    border-bottom: 3px solid #4361ee;
    padding-bottom: .5rem; margin-bottom: 1.5rem;
  }}
  .stats {{
    background: linear-gradient(135deg, #e8eaf6, #f3e5f5);
    padding: 1rem 1.5rem; border-radius: 10px;
    margin-bottom: 2rem; font-size: .95rem;
    box-shadow: 0 2px 6px rgba(0,0,0,.08);
    display: flex; gap: 2rem; flex-wrap: wrap;
  }}
  .stats .stat-item {{ font-weight: 600; }}
  .stats .stat-value {{ color: #4361ee; }}
  .toc {{
    background: #fff; padding: 1.5rem 2rem; border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,.07); margin-bottom: 2rem;
  }}
  .toc h2 {{ font-size: 1.2rem; margin-bottom: .8rem; color: #4361ee; }}
  .toc ul {{ list-style: none; columns: 2; column-gap: 2rem; }}
  .toc li {{ padding: .35rem 0; }}
  .toc a {{
    color: #4361ee; text-decoration: none; font-weight: 500;
    transition: color .2s;
  }}
  .toc a:hover {{ color: #3a0ca3; text-decoration: underline; }}
  .toc .count {{ color: #888; font-size: .85rem; margin-left: .3rem; }}
  .topic-section {{
    background: #fff; border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,.07);
    padding: 1.5rem 2rem; margin-bottom: 1.5rem;
    border-left: 5px solid #ccc;
  }}
  .topic-header {{
    font-size: 1.3rem; color: #1a1a2e;
    padding-left: .5rem; margin-bottom: 1rem;
    display: flex; align-items: center; gap: .5rem;
  }}
  .topic-header .badge {{
    background: #4361ee; color: #fff; font-size: .75rem;
    padding: 2px 10px; border-radius: 12px;
  }}
  .article {{
    border-bottom: 1px solid #eee; padding: 1rem 0;
    transition: background .15s;
  }}
  .article:last-child {{ border-bottom: none; }}
  .article:hover {{ background: #fafbff; border-radius: 6px; padding-left: .5rem; }}
  .article-title {{ font-size: 1.05rem; font-weight: 600; }}
  .article-title a {{ color: #4361ee; text-decoration: none; }}
  .article-title a:hover {{ text-decoration: underline; color: #3a0ca3; }}
  .article-meta {{
    font-size: .85rem; color: #666; margin: .3rem 0 .6rem;
    display: flex; gap: .8rem; flex-wrap: wrap; align-items: center;
  }}
  .article-meta .date {{ font-weight: 500; }}
  .tag {{
    background: #e8eaf6; padding: 2px 10px; border-radius: 12px;
    font-size: .78rem; color: #4361ee; display: inline-block;
  }}
  .article-summary {{
    font-size: .93rem; line-height: 1.6; color: #333;
    max-width: 900px;
  }}
  .article-summary b {{ color: #1a1a2e; }}
  .article-summary ul {{ margin: .4rem 0 .4rem 1.2rem; }}
  .article-summary li {{ margin-bottom: .2rem; }}
  .back-link {{
    display: inline-block; margin-top: 1rem; color: #4361ee;
    font-size: .85rem; text-decoration: none; font-weight: 500;
  }}
  .back-link:hover {{ text-decoration: underline; }}
  .dup-badge {{
    background: #fff3cd; color: #856404; padding: 2px 8px;
    border-radius: 4px; font-size: .78rem; font-weight: 500;
  }}
  footer {{
    text-align: center; color: #999; font-size: .8rem;
    margin-top: 3rem; padding-top: 1rem;
    border-top: 1px solid #e0e0e0;
  }}
</style></head><body>
<h1>Blog Notifications &mdash; {date}</h1>
<div class="stats" id="stats"></div>
<nav class="toc" id="toc"><h2>Topics</h2><ul id="toc-list"></ul></nav>
<div id="content"></div>
<footer>Generated by Blog Notification Pipeline</footer>
</body></html>""".format(date=today)

with open(html_path, "w", encoding="utf-8") as f:
    f.write(html)

# ── Initialize session state ─────────────────────────────────────────────
session = {
    "date": today,
    "xlsx_path": xlsx_path,
    "html_path": html_path,
    "emails": [],
    "processed_titles": {},  # title -> final_url for session dedup
}
json.dump(session, open(session_file, "w", encoding="utf-8"), indent=2, ensure_ascii=False)

# ── Output result ────────────────────────────────────────────────────────
result = {
    "xlsx_path": xlsx_path,
    "html_path": html_path,
    "session_file": session_file,
    "base_name": base_name,
}
print(json.dumps(result, indent=2))
